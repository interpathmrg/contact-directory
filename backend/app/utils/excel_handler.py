"""
Utilidades para parsear y generar archivos Excel/CSV de contactos.
Sin dependencias de BD — solo transformaciones de datos.
"""
from __future__ import annotations

import csv
import io

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Definición maestra de columnas: (field_name, label_xlsx, required)
COLUMNS: list[tuple[str, str, bool]] = [
    ("nombre",                   "Nombre",                   True),
    ("apellido",                 "Apellido",                 True),
    ("empresa",                  "Empresa",                  False),
    ("cargo",                    "Cargo",                    False),
    ("puesto",                   "Puesto",                   False),
    ("direccion",                "Dirección",                False),
    ("telefono",                 "Teléfono",                 False),
    ("celular",                  "Celular",                  False),
    ("email",                    "Email",                    False),
    ("nombre_contacto_interno",  "Nombre Contacto Interno",  False),
    ("email_contacto_interno",   "Email Contacto Interno",   False),
    ("telefono_contacto_interno","Teléfono Contacto Interno",False),
    ("nota",                     "Nota",                     False),
    ("sociedad",                 "Sociedad",                 False),
]

# Mapeo inverso: label → field_name  (para parsear archivos del usuario)
_LABEL_TO_FIELD: dict[str, str] = {label.lower(): field for field, label, _ in COLUMNS}
_FIELD_TO_LABEL: dict[str, str] = {field: label for field, label, _ in COLUMNS}

_HEADER_COLOR = "1d4ed8"
_REQUIRED_COLOR = "1e40af"


def _to_str(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def generate_template(society_names: list[str] | None = None) -> bytes:
    """Genera el Excel plantilla para importación con headers estilizados."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Contactos"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    req_fill = PatternFill(start_color=_REQUIRED_COLOR, end_color=_REQUIRED_COLOR, fill_type="solid")
    opt_fill = PatternFill(start_color=_HEADER_COLOR, end_color=_HEADER_COLOR, fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 20

    for col_idx, (field, label, required) in enumerate(COLUMNS, 1):
        header_text = f"{label} *" if required else label
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.font = header_font
        cell.fill = req_fill if required else opt_fill
        cell.alignment = center
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(header_text) + 4, 18)

    # Fila de ejemplo
    example = {
        "nombre": "Juan",
        "apellido": "Pérez García",
        "empresa": "Empresa Ejemplo S.A.",
        "cargo": "Gerente",
        "email": "juan.perez@ejemplo.com",
        "telefono": "809-555-1234",
        "celular": "829-555-5678",
        "sociedad": society_names[0] if society_names else "EGE HAINA",
    }
    for col_idx, (field, _, _r) in enumerate(COLUMNS, 1):
        ws.cell(row=2, column=col_idx, value=example.get(field, ""))

    # Instrucciones en hoja separada
    ws_info = wb.create_sheet("Instrucciones")
    ws_info["A1"] = "INSTRUCCIONES DE IMPORTACIÓN"
    ws_info["A1"].font = Font(bold=True, size=13)
    ws_info["A3"] = "Campos obligatorios:"
    ws_info["A4"] = "  • Nombre *"
    ws_info["A5"] = "  • Apellido *"
    ws_info["A7"] = "Campo Sociedad — valores válidos:"
    for i, s in enumerate(society_names or ["EGE HAINA", "SIBA", "Trelia"], 8):
        ws_info[f"A{i}"] = f"  • {s}"
    ws_info["A20"] = "• Emails duplicados en el archivo o ya existentes en el sistema serán marcados como error."
    ws_info["A21"] = "• Las filas vacías son ignoradas automáticamente."
    ws_info.column_dimensions["A"].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_excel(content: bytes) -> list[dict[str, str]]:
    """Parsea un .xlsx y retorna lista de dicts con las claves en field_name."""
    buf = io.BytesIO(content)
    wb = load_workbook(buf, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    raw_headers = [_to_str(h).lower() for h in rows[0]]
    col_map: dict[int, str] = {}
    for i, raw_h in enumerate(raw_headers):
        # Quitar el asterisco de campos requeridos ("nombre *" → "nombre")
        clean = raw_h.rstrip(" *").strip()
        field = _LABEL_TO_FIELD.get(clean) or _LABEL_TO_FIELD.get(raw_h)
        if field:
            col_map[i] = field

    result: list[dict[str, str]] = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        row_dict = {field: "" for field, _, _ in COLUMNS}
        for i, field in col_map.items():
            row_dict[field] = _to_str(row[i] if i < len(row) else None)
        result.append(row_dict)

    return result


def parse_csv(content: bytes) -> list[dict[str, str]]:
    """Parsea un .csv (auto-detecta encoding) y retorna lista de dicts."""
    text: str = ""
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue

    reader = csv.DictReader(io.StringIO(text))
    result: list[dict[str, str]] = []
    for raw_row in reader:
        if not any(raw_row.values()):
            continue
        row_dict = {field: "" for field, _, _ in COLUMNS}
        for raw_key, raw_val in raw_row.items():
            clean_key = raw_key.strip().lower().rstrip(" *")
            field = _LABEL_TO_FIELD.get(clean_key) or _LABEL_TO_FIELD.get(raw_key.strip().lower())
            if field:
                row_dict[field] = (raw_val or "").strip()
        result.append(row_dict)
    return result


def contacts_to_excel(contacts: list[dict]) -> bytes:
    """Genera un Excel con la lista de contactos para exportación."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Contactos"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=_HEADER_COLOR, end_color=_HEADER_COLOR, fill_type="solid")

    for col_idx, (field, label, _) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(label) + 4, 18)

    for row_idx, contact in enumerate(contacts, 2):
        for col_idx, (field, _, _r) in enumerate(COLUMNS, 1):
            ws.cell(row=row_idx, column=col_idx, value=contact.get(field, ""))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def contacts_to_csv(contacts: list[dict]) -> bytes:
    """Genera un CSV UTF-8 BOM (compatible con Excel) con la lista de contactos."""
    labels = [label for _, label, _ in COLUMNS]
    fields = [field for field, _, _ in COLUMNS]

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(labels)
    for contact in contacts:
        writer.writerow([contact.get(field, "") for field in fields])

    return buf.getvalue().encode("utf-8-sig")
